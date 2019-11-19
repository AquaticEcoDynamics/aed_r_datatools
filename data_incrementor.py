import pandas as pd
from xlrd import open_workbook
import xlrd
import os
from xlutils.copy import copy
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.conftest import Workbook
import datetime
def data_incrementor(dir, file1, file2):
    wb = xlrd.open_workbook(file1)
    sheet = wb.sheet_by_index(0)
    wb_wt = open_workbook(file2)
    first_workbook = openpyxl.load_workbook(file2)
    new_workbook = openpyxl.Workbook()



    new_index = 0
    added_rows = 0
    for s in wb_wt.sheets():
        sheetname = new_workbook.active
        sheetname1 = first_workbook[s.name]
        for i in range(1,sheet.nrows):

            current_date_tuple = xlrd.xldate_as_tuple(sheet.cell_value(i,2),wb.datemode)
            if i >= 3:
                start_date = datetime.datetime(current_date_tuple[0],current_date_tuple[1],current_date_tuple[2])
                end_date = datetime.datetime(previous_date_tuple[0],previous_date_tuple[1],previous_date_tuple[2])
                delta_days = end_date - start_date
                seconds_total = delta_days.total_seconds()
                if seconds_total > 24*3600:
                    date = start_date
                    print(True)
                    while True:
                        date += datetime.timedelta(days = 1)
                        if date.day == end_date.day and date.month == end_date.month and date.year == end_date.year:
                            break
                        added_rows += 1
            previous_date_tuple = current_date_tuple
        for i in range(1,sheet.nrows):
            current_date_tuple = xlrd.xldate_as_tuple(sheet.cell_value(i,2),wb.datemode)
            if i == 1:
                sheetname.cell(i,0+1).value = sheetname1.cell(i,1).value
                sheetname.cell(i,4+1).value = sheetname1.cell(i,5).value
                sheetname.cell(i,5+1).value = sheetname1.cell(i,6).value
                sheetname.cell(i,1+1).value = sheetname1.cell(i,2).value
                sheetname.cell(i,2+1).value = sheetname1.cell(i,3).value
                sheetname.cell(i,3+1).value = sheetname1.cell(i,4).value
            if i == 2:
                sheetname.cell(sheet.nrows + added_rows - new_index-1,0+1).value = sheetname1.cell(i,1).value
                sheetname.cell(sheet.nrows + added_rows - new_index-1,4+1).value = sheetname1.cell(i,5).value
                sheetname.cell(sheet.nrows + added_rows - new_index-1,5+1).value = sheetname1.cell(i,6).value
                sheetname.cell(sheet.nrows + added_rows - new_index-1,1+1).value = sheetname1.cell(i,2).value
                sheetname.cell(sheet.nrows + added_rows - new_index-1,2+1).value = str(current_date_tuple[1])+ '/' +str(current_date_tuple[2])+ '/'+str(current_date_tuple[0])
                sheetname.cell(sheet.nrows + added_rows - new_index-1,3+1).value = sheetname1.cell(i,4).value
                new_index += 1
            if i == 3:
                id_2 = sheetname1.cell(i,1).value
                id_1 = sheetname1.cell(i-1,1).value
                id_2_integ = int(id_2[1:len(id_2)])
                id_1_integ = int(id_1[1:len(id_1)])
                new_id = id_2_integ
                if id_2_integ <id_1_integ:
                    id_flag = -1
                elif id_2_integ >id_1_integ:
                    id_flag = 1
                else:
                    id_flag = 0
            if i >= 3:
                start_date = datetime.datetime(current_date_tuple[0],current_date_tuple[1],current_date_tuple[2])
                end_date = datetime.datetime(previous_date_tuple[0],previous_date_tuple[1],previous_date_tuple[2])
                delta_days = end_date - start_date
                seconds_total = delta_days.total_seconds()
                if seconds_total > 24*3600:
                    date = end_date
                    print(i)
                    while True:
                        date -= datetime.timedelta(days = 1)
                        if date.day == start_date.day and date.month == start_date.month and date.year == start_date.year:
                            break
                        if id_flag == -1:
                            new_id -= 1
                        elif id_flag == 1:
                            new_id += 1
                        mix_id = id_2[0] + str(new_id)
                        sheetname.cell(sheet.nrows + added_rows - new_index-1,0+1).value = mix_id
                        sheetname.cell(sheet.nrows + added_rows - new_index-1,4+1).value = 'B'
                        sheetname.cell(sheet.nrows + added_rows - new_index-1,5+1).value = 'WC'
                        sheetname.cell(sheet.nrows + added_rows - new_index-1,1+1).value = 1
                        sheetname.cell(sheet.nrows + added_rows - new_index-1,2+1).value = str(date.month) + '/' + str(date.day) + '/' + str(date.year)
                        sheetname.cell(sheet.nrows + added_rows - new_index-1,3+1).value = ''
#                        added_rows += 1
                        new_index += 1
#                    new_index += 1

                else:
                    mean_value = sheet.cell_value(i,5)
                    date_string = str(current_date_tuple[1])+ '/' +str(current_date_tuple[2])+ '/'+str(current_date_tuple[0])
                    if id_flag == -1:
                        new_id -= 1
                    elif id_flag == 1:
                        new_id += 1
                    mix_id = id_2[0] + str(new_id)
                    sheetname.cell(sheet.nrows + added_rows - new_index-1,0+1).value = mix_id
                    sheetname.cell(sheet.nrows + added_rows - new_index-1,4+1).value = 'B'
                    sheetname.cell(sheet.nrows + added_rows - new_index-1,5+1).value = 'WC'
                    sheetname.cell(sheet.nrows + added_rows - new_index-1,1+1).value = 1
                    sheetname.cell(sheet.nrows + added_rows - new_index-1,2+1).value = date_string
                    sheetname.cell(sheet.nrows + added_rows - new_index-1,3+1).value = mean_value
                    new_index += 1

            previous_date_tuple = current_date_tuple
    new_workbook.save( dir+'\Pinjarrah'+ os.path.splitext(file2)[-1])

def main():
    root_dir = 'Alilou'
    index = 0
    for subdir, dirs, files in os.walk(root_dir):
        for files in files:
            if os.path.join(subdir,files).__contains__('.xlsx') and any(char.isdigit() for char in files):
                print(os.path.join(subdir,files))
                data_incrementor(subdir,os.path.join(subdir,files),'Pinjarra1.xlsx')
                index  += 1
                if index == 1:
                    break


if __name__ == '__main__':
    main()